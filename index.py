from flask import Flask, request, render_template, send_file
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

def calculate_offset(number):
    return 15 + ((number - 1) // 2) * 12

def shift_letter(column, shift):
    def col_to_num(col):
        num = 0
        for char in col:
            num = num * 26 + (ord(char) - ord('A') + 1)
        return num

    def num_to_col(num):
        col = ""
        while num > 0:
            num -= 1
            col = chr(num % 26 + ord('A')) + col
            num //= 26
        return col

    column = column.upper()
    new_num = col_to_num(column) + shift
    return num_to_col(new_num)

def map_location_to_cell(location):
    rack, shelf, bin = location.split('-')
    rack = int(rack)
    rack_is_odd = rack % 2 != 0
    odd_map = {'a': 1, 'b': 2, 'c': 3, 'd': 4, 'e': 5}
    even_map = {'e': 6, 'd': 7, 'c': 8, 'b': 9, 'a': 10}
    
    excelRow = calculate_offset(rack) + (odd_map[shelf.lower()] if rack_is_odd else even_map[shelf.lower()])
    excelColumn = int(bin) + (7 if rack < 51 else 17)
    excelColumn = shift_letter('A', excelColumn - 1)
    
    return f"{excelColumn}{str(excelRow)}"

def highlight_warehouse(location, sku, file_path, output_path):
    wb = load_workbook(file_path)
    ws = wb.active
    
    sku_colors = {
        'sku1': PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid'),
        'sku2': PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    }
    
    cell = map_location_to_cell(location)
    ws[cell].fill = sku_colors.get(sku, PatternFill(fill_type=None))
    ws[cell] = "Highlighted"
    
    wb.save(output_path)

app = Flask(__name__)
UPLOAD_FOLDER = "uploads"
OUTPUT_FOLDER = "output"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

@app.route('/')
def index():
    return render_template('viewAllocations.html')

@app.route('/process', methods=['POST'])
def process():
    if 'file' not in request.files:
        return "No file uploaded", 400
    
    file = request.files['file']
    location = request.form['location']
    sku = request.form['sku']
    
    file_path = os.path.join(UPLOAD_FOLDER, file.filename)
    output_path = os.path.join(OUTPUT_FOLDER, f"highlighted_{file.filename}")
    
    file.save(file_path)
    highlight_warehouse(location, sku, file_path, output_path)
    
    return send_file(output_path, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)

import pandas as pd
import string
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import colorsys

def generate_distinct_pastel_colors(n):
    """Generate `n` distinct pastel colors using HSL for better separation."""
    colors = []
    for i in range(n):
        hue = i / n  # Evenly spaced hues
        r, g, b = colorsys.hls_to_rgb(hue, 0.8, 0.5)  # Pastel-like brightness
        colors.append(f"{int(r * 255):02X}{int(g * 255):02X}{int(b * 255):02X}")
    return colors

def assign_sku_colors(skus):
    """Dynamically assign distinct pastel colors to SKUs."""
    distinct_colors = generate_distinct_pastel_colors(len(skus))
    return {sku: PatternFill(start_color=color, end_color=color, fill_type='solid')
            for sku, color in zip(skus, distinct_colors)}


def calculate_offset(number):
    return 15 + ((number - 1) // 2) * 12

def shift_letter(column, shift):
    def col_to_num(col):
        """Convert Excel column label (e.g., 'A', 'Z', 'AA') to a number (e.g., 1, 26, 27)."""
        num = 0
        for char in col:
            num = num * 26 + (ord(char) - ord('A') + 1)
        return num

    def num_to_col(num):
        """Convert a number back to an Excel column label."""
        col = ""
        while num > 0:
            num -= 1  # Adjust for 1-based indexing
            col = chr(num % 26 + ord('A')) + col
            num //= 26
        return col

    column = column.upper()  # Ensure uppercase
    new_num = col_to_num(column) + shift  # Shift column number
    return num_to_col(new_num)  # Convert back to column label

def map_location_to_cell(location):
    """
    Converts a warehouse location (e.g., "1-A-1") to an Excel cell (e.g., "H16").
    """
    rack, shelf, bin = location.split('-')
    rack = int(rack)
    rack_is_odd = rack % 2 != 0
    odd_map = {
        'a': 1,
        'a1': 1,
        'a2': 1,
        'b': 2,
        'c': 3,
        'd': 4,
        'e': 5
    }
    even_map = {
        'e': 6,
        'd': 7,
        'c': 8,
        'b': 9,
        'a': 10,
        'a1': 10,
        'a2': 10,
    }
    if rack_is_odd:
        excelRow = calculate_offset(rack) + odd_map[shelf.lower()]
    else:
        excelRow = calculate_offset(rack) + even_map[shelf.lower()]
# if rack is an even number, calculate_offset(rack_number), then add {{mapShelfEven(shelf_number)}}
# if rack is an odd number, calculate_offset(rack_number), then add {{mapShelfOdd(shelf_number)}}

    print(bin)
    excelColumn = int(bin) + 7
    print(excelColumn)
    excelColumn = shift_letter('A', excelColumn - 1)
    print(excelColumn)
    
    print(f"{excelColumn}{str(excelRow)}")
    return f"{excelColumn}{str(excelRow)}"

def highlight_warehouse(locations, file_path, output_path):
    """
    Highlights warehouse locations in an Excel file.
    """
    wb = load_workbook(file_path)
    ws = wb.active
    print(ws)
    # Define SKU-based colors
    sku_colors = assign_sku_colors([info['sku'] for info in locations.values()])
    
    for location, info in locations.items():
        sku = info['sku']
        quantity = info['quantity']
        cell = map_location_to_cell(location)
        ws[cell].fill = sku_colors.get(sku, PatternFill(fill_type=None))  # Default: No color
        ws[cell] = sku + ': ' + str(quantity)  # Display quantity in the cell
    
    # Save the updated file
    wb.save(output_path)


if __name__ == '__main__':
# Example usage
    file_path ="C:\\Users\\Owner\\Downloads\\WM warehouse.xlsx"
    output_path = "C:\\Users\\Owner\\Downloads\\WM1warehouse.xlsx"
    locations_to_highlight = {"1-A-1": "sku1", "10-E-2": "sku2", "25-D-30": "sku1"}
    locations =  {
            "53-d-45": { "sku": "OLN-ERGOACE-CRM", "quantity": 8 },
            "53-d-44": { "sku": "OLN-ERGOACE-CRM", "quantity": 8 },
            "53-c-44": { "sku": "OLN-ERGOACE-CRM", "quantity": 7 },
            "53-d-40": { "sku": "OLN-ERGOACE-CRM", "quantity": 8 },
            "53-e-30": { "sku": "OLN-ERGOACE-CRM", "quantity": 8 },
            "52-a-03": { "sku": "OLN-ERGOACE-CRM", "quantity": 7 },
            "52-a1-15": { "sku": "OLN-ERGOACE-CRM", "quantity": 4 }
            }

    highlight_warehouse(locations, "./WM_warehouse.xlsx", "../public/WM1warehouse.xlsx")